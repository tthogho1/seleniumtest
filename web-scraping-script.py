import os
import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import openpyxl
import time

write_count = 0

def find_empty_column(ws):
    """
    最初の空いている列を見つける関数
    
    Args:
        ws (openpyxl.worksheet.Worksheet): 検索対象のワークシート
    
    Returns:
        int: 最初の空いている列の番号
    """
    max_row = ws.max_row
    for col in range(1, ws.max_column + 2):  # +2 は新しい列を確保するため
        is_column_empty = True
        for row in range(1, max_row + 1):
            if ws.cell(row=row, column=col).value is not None:
                is_column_empty = False
                break
        if is_column_empty:
            return col
    return ws.max_column + 1

def scrape_input_tags(driver):
    """
    開いているブラウザウィンドウからinputタグを取得し、Excelに出力する関数
    
    Args:
        driver (webdriver): Seleniumのwebドライバー
    """
    global write_count
    
    try:
        # ページの読み込みを待つ
        time.sleep(3)  # 必要に応じて調整
        
        # 現在のURLを取得
        current_url = driver.current_url
        print(f"現在のURL: {current_url}")
        
        # HTMLを取得
        html = driver.page_source
        
        # BeautifulSoupでHTMLを解析
        soup = BeautifulSoup(html, 'html.parser')
        
        # inputタグのtype、name、valueを取得
        input_data = []
        for input_tag in soup.find_all('input'):
            input_type = input_tag.get('type', 'N/A')
            input_name = input_tag.get('name', 'N/A')
            input_value = input_tag.get('value', 'N/A')
            input_data.append({
                'type': input_type, 
                'name': input_name, 
                'value': input_value
            })
        
        # Excel処理
        excel_file = 'input_tags.xlsx'
        
        # ファイルが存在しない場合は新規作成
        if not os.path.exists(excel_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                return
            wb.save(excel_file)
            
        # ワークブックを開く
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        if ws is None:
            # エラーを表示
            print(f'Excelファイル {excel_file} を開けません。')
            return
        
        
        # 空いている列を見つける
        start_col = find_empty_column(ws)        
        
        ## 1から始める
        write_count = write_count + 1
        # ヘッダーを書き込む
        ws.cell(row=1, column=start_col, value='type'+ str(write_count))
        ws.cell(row=1, column=start_col + 1, value='name' + str(write_count))
        ws.cell(row=1, column=start_col + 2, value='value' + str(write_count))
        #wb.save(excel_file)s
        
        # データを書き込む
        for idx, item in enumerate(input_data, start=2):  # ヘッダー行の次から
            ws.cell(row=idx, column=start_col, value=item['type'])
            ws.cell(row=idx, column=start_col + 1, value=item['name'])
            ws.cell(row=idx, column=start_col + 2, value=item['value'])
        
        # 保存
        wb.save(excel_file)
        print(f'Excelファイル {excel_file} に出力しました。')
    
    except Exception as e:
        print(f"スクレイピング中にエラーが発生しました: {e}")

def main():
    print("Input Tag Scraper")
    print("----------------")
    print("1. Chromeを開いてください")
    print("2. 目的のウェブページを開いてください")
    print("3. このプログラムに戻り、'x'を押してスクレイピングを実行")
    print("4. 'q'で終了")
    
    # Selenium WebDriverの設定（Chrome）
    driver = webdriver.Chrome()  # 必要に応じてChromeDriverのパスを指定
        
    try:
        while True:
            # ユーザー入力を待つ
            user_input = input("入力してください: ").lower()
            
            if user_input == 'x':
                # 現在開いているブラウザウィンドウをスクレイピング
                scrape_input_tags(driver)
            
            elif user_input == 'q':
                print("プログラムを終了します。")
                break
            
            else:
                print("無効な入力です。'x'または'q'を入力してください。")
    
    finally:
        # ブラウザを閉じる
        driver.quit()

if __name__ == "__main__":
    main()
